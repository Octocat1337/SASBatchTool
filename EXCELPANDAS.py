import openpyxl
import os
from time import sleep
import warnings

class EXCELHandler2:
    current_folder = ''
    tracker_folder = ''
    tracker_file = ''

    # Global Vars
    sdtm_sheet_name = 'SDTM Dataset'
    tlf_sheet_name = 'TLF'

    topline_pos = 0
    intext_pos = 0
    combine_pos = 0
    program_pos = 0
    qcprogram_pos = 0

    is_qc = False

    def __init__(self, folder='', dummy=False):
        """
            :param folder: current folder path
        """
        self.current_folder = folder

        if dummy:
            return

        if '\\qc\\' in self.current_folder or '/qc/' in self.current_folder:
            self.is_qc = True

        self.tracker_folder = self.get_tracker_folder(path=self.current_folder,testenv=True)

        # get tracker excel file
        file_list_tmp = os.listdir(self.tracker_folder)
        foundFile=False
        for item in file_list_tmp:
            # potential bug: multiple tracker files may be present
            # current solution: get the 1st one
            if 'tracker.xls' in item:
                self.tracker_file = item
                foundFile=True
                break
        if not foundFile: print('Tracker file not found')
        self.tracker_file_path = os.path.join(self.tracker_folder,self.tracker_file)


    def get_filelist_dummy(self, type='', root=None):
        root.event_generate("<<event2>>", when='tail', state=1)
        sleep(5)
        root.event_generate("<<event2>>", when='tail', state=2)
        list = []
        return list

    def get_filelist(self, tlf_type='', root=None):
        root.event_generate("<<event2>>", when='tail', state=1)

        file_path = self.tracker_file_path
        return_list=[]
        sheet_name = 'TLF'

        try:
            # pd.read_excel reads the first row as column headers by default
            # suppress openpyxl user warning: excel's data validation for droplist
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning)
                # df = pd.read_excel(file_path, sheet_name=sheet_name)
                wb = openpyxl.load_workbook(file_path, data_only=True)
                if sheet_name not in wb.sheetnames:
                    print("TLF sheet not found")
                    return []

            sheet = wb["TLF"]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

            tlf_type_idx = None
            program_idx = None
            for i, column_name in enumerate(header_row):
                if column_name:
                    col_lower = str(column_name).strip().lower()
                    if col_lower == tlf_type:
                        tlf_type_idx = i
                    elif col_lower == "program" and not self.is_qc:
                        program_idx = i
                    elif col_lower == "qc program" and self.is_qc:
                        program_idx = i

            if tlf_type_idx is None or program_idx is None:
                print("[ERR] in get_filelist, did not find in header: "+tlf_type)
                return []

            # Iterate through rows starting from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                tlf_type_idx_val = str(row[tlf_type_idx]).strip().upper() if row[tlf_type_idx] else ""

                # Check if the "topline" value is "Y"
                if tlf_type_idx_val == "Y":
                    if row[program_idx] is not None:
                        return_list.append(row[program_idx]+'.sas')

        except FileNotFoundError:
            print(f"Error: The file '{file_path}' was not found.")
        except Exception as e:
            print(f"An error occurred: {e}")

        # print("-----Exit ExcelHandler2-----")
        root.event_generate("<<event2>>", when='tail', state=2)
        return return_list

    def get_tracker_folder(self, path='',target_folder='program',testenv=False):
        '''
        a\b\c\d\dryrun1\program\tlf -> a\b\c\d\dryrun1
        :return: dryrun1 level folder
        '''
        normpath = os.path.normpath(path)
        parts = normpath.split(os.sep)  # Split the path into parts using the OS separator
        try:
            index = parts.index(target_folder)
            base_path = ''
            if self.is_qc:
                if testenv:
                    base_path = os.path.join(parts[0]+os.sep,*parts[:index -1],'document','tracker')
                else:
                    base_path = os.path.join('Z:\\',*parts[1:index-1],'document','tracker')
            else:
                if testenv:
                    # base_path = os.path.join(*parts[:index],'document','tracker')
                    base_path = os.path.join(parts[0]+os.sep,*parts[1:index],'document','tracker')
                    print("base_path: " + base_path)
                else:
                    base_path = os.path.join('Z:\\',*parts[1:index],'document','tracker')
            return base_path
        except ValueError:
            print("ERROR in get_tracker_folder function!")
            return ''  # Target folder not found
